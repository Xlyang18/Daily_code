import openpyxl
import re
from collections import Counter

date_one = []
date_Time = []
g_coordinate = []
b_coordinate = []
n_coordinate = []


def question_2():
    test_a = ['D', '设计内容圈层', 'F', '品牌、潮牌、设计师', 'G', '购买动机',
              'H', '购买渠道', 'I', '购买风格', 'J', '内容发布风格', 'K', '信息获取渠道',
              'L', '圈子、文化认同感', 'R', '认证类型', 'V', '性别', 'W', '地域',
              'END']
    num_a = 0
    while test_a[num_a] != 'END':
        column_date = ws[test_a[num_a]]
        print("影响因素：", test_a[num_a + 1])
        for i in range(1, num):
            date_one.append(column_date[i].value)

        statistics = dict(Counter(date_one))
        print("重复出现的购买动因有：", [key for key, value in statistics.items() if value > 1])
        print("统计出现的次数:")
        content = {key: value for key, value in statistics.items() if value > 1}
        content = str(content)

        # 替换字符串，换行处理
        x, n = re.subn(", ", "\n", content, )
        print(x)
        date_one.clear()
        num_a = num_a + 2
        print("-------End of Influence Factor-------\n")


def question_1():
    num_g = 0
    num_b = 0
    num_n = 0
    column_date = ws["V2:V2333"]
    # print(column_date.value)
    for one in column_date:
        for cell in one:
            print(cell.coordinate, cell.value)
            # 获取筛选后的序号
            cell_coordinate = str(cell.coordinate)
            cell_coordinate = cell_coordinate[1:]
            print(">>>>", cell_coordinate)
            one_cell = str(cell.value)
            if one_cell == "女":
                # !!!
                num_g = num_g + 1

                cell_coordinate_g = str(cell.coordinate)
                cell_coordinate_g = cell_coordinate_g[1:]
                print("女-行号>>>>", cell_coordinate_g)
                g_coordinate.append(cell_coordinate_g)

            elif one_cell == "男":
                num_b = num_b + 1

                cell_coordinate_b = str(cell.coordinate)
                cell_coordinate_b = cell_coordinate_b[1:]
                print("男-行号>>>>", cell_coordinate_b)
                b_coordinate.append(cell_coordinate_b)

            elif one_cell == "未知":
                num_n = num_n + 1

                cell_coordinate_n = str(cell.coordinate)
                cell_coordinate_n = cell_coordinate_n[1:]
                print("未知-行号>>>>", cell_coordinate_n)
                n_coordinate.append(cell_coordinate_n)

        print("-----End of Row-----")
        print("num_g:{}| num_b:{}| num_n:{}".format(num_g, num_b, num_n))

    # 将字符串数组转换为数字数组
    new_g_corrdinate = []
    new_b_corrdinate = []
    new_n_corrdinate = []
    for n_g in g_coordinate:
        new_g_corrdinate.append(int(n_g))
    for n_b in b_coordinate:
        new_b_corrdinate.append(int(n_b))
    for n_n in n_coordinate:
        new_n_corrdinate.append(int(n_n))
    print("女行号统计有：{}\n男行号统计有：{}\n未知行号统计有：{}".format(new_g_corrdinate, new_b_corrdinate, new_n_corrdinate))

    test_a = ['D', '设计内容圈层', 'F', '品牌、潮牌、设计师', 'G', '购买动机',
              'H', '购买渠道', 'I', '购买风格', 'J', '内容发布风格', 'K', '信息获取渠道',
              'L', '圈子、文化认同感', 'R', '认证类型', 'END']
    num_a = 0
    while test_a[num_a] != 'END':
        column_date = ws[test_a[num_a]]
        print("影响因素：", test_a[num_a + 1])

        # 此处的new_g_corrdinate 可改变数组参数值，查看其他用户群体的数据
        # new_g_corrdinate、new_b_corrdinate、new_n_corrdinate
        for i in new_g_corrdinate:
            date_one.append(column_date[i].value)

        statistics = dict(Counter(date_one))
        print("重复出现的购买动因有：", [key for key, value in statistics.items() if value > 1])
        print("统计出现的次数:")
        content = {key: value for key, value in statistics.items() if value > 1}
        content = str(content)

        # 替换字符串，换行处理
        x, n = re.subn(", ", "\n", content, )
        print(x)
        date_one.clear()
        num_a = num_a + 2
        print("-------End of Influence Factor-------\n")


    # 打印A1～C3单元格内所有单元格元素的内容
    # test_a = ['D', '设计内容圈层', 'F', '品牌、潮牌、设计师', 'G', '购买动机',
    #           'H', '购买渠道', 'I', '购买风格', 'J', '内容发布风格', 'K', '信息获取渠道',
    #           'L', '圈子、文化认同感', 'R', '认证类型', 'V', '性别', 'W', '地域',
    #           'END']
    # i = 0
    # for i in range(1, num):
    #     if ws["V"] == "男":
    #         date_one.append(column_date[i].value)


# 开始
global num
wb = openpyxl.load_workbook("C题 附件.xlsx")
ws = wb.active
print("操作的对应表单为：", ws, )
print("对应表单的长度为{}*{}".format(ws.max_row, ws.max_column))
row_long = ws.max_row
num = int(row_long)

question_1()

